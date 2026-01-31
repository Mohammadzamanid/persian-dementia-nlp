# -*- coding: utf-8 -*-
"""
Recompute + update ALL sheets in:
  paper1_final_outputs-mpnet2_WITH_MERGED_FEATURES.xlsx

Adds:
  coh_composite_mean = mean(coh_local_mean, coh_global_mean)

Rebuilds these sheets:
  - UnifiedDataset
  - MergeDiagnostics (kept as-is from input workbook if present)
  - Missingness
  - Desc_Demographics
  - Sex_Distribution
  - Desc_Language
  - ANCOVA_All
  - PostHoc_Contrasts

Models:
  ModelA_no_LII: outcome ~ group + age + education + sex
  ModelB_with_LII: outcome ~ group + age + education + sex + lii_mean
  (ModelB is NOT run for outcome=lii_mean)

ANCOVA uses statsmodels anova_lm(typ=2).  :contentReference[oaicite:0]{index=0}
Holm adjustment is implemented to match the workbook (within-outcome, 3 contrasts). :contentReference[oaicite:1]{index=1}
"""

from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

import statsmodels.formula.api as smf
from statsmodels.stats.anova import anova_lm
from scipy.stats import t as tdist


# ----------------- CONFIG -----------------
IN_PATH  = Path(r"C:\Users\Fatiima\Desktop\voices\Thesis_2\persian_norm\paper1_final_outputs-mpnet2_WITH_MERGED_FEATURES4.xlsx")
OUT_PATH = Path(r"C:\Users\Fatiima\Desktop\voices\Thesis_2\persian_norm\paper1_final_outputs-mpnet2_WITH_MERGED_FEATURES_updated_all_sheets-2.xlsx")

# Sheet names (must match your workbook)
SHEET_UNIFIED = "UnifiedDataset"
SHEET_DIAG    = "MergeDiagnostics"
SHEET_MISS    = "Missingness"
SHEET_DEM     = "Desc_Demographics"
SHEET_SEX     = "Sex_Distribution"
SHEET_DESC    = "Desc_Language"
SHEET_ANCOVA  = "ANCOVA_All"
SHEET_POSTHOC = "PostHoc_Contrasts"

GROUP_ORDER = ["Control", "MCI", "Mild_AD"]
ALL_LABEL   = "All"

# Outcome order as it appears in your output workbook
OUTCOME_ORDER = [
    "lii_mean",
    "silence_to_audio_ratio",
    "avg_branching_factor",
    "mls_tokens",
    "tree_depth_mean",
    "dependency_length_mean",
    "noun_to_pronoun_ratio",
    "noun_to_verb_ratio",
    "graph_density_noloop",
    "graph_lscc_ratio",
    "graph_self_loops",
    "coh_local_mean",
    "coh_global_mean",
    "coh_composite_mean",
    "pid_repeat_ratio",
    "graph_repeated_edges_per100",
    "graph_tokens",
    "graph_nodes",
    "graph_edges_unique",
    "graph_edges_unique_noloop",
    "graph_lscc_size",
    "graph_lcc_size",
    "graph_lcc_ratio",
    "graph_avg_degree",
    "coh_n_utts",
    "coh_local_sd",
    "coh_global_sd",
    "pid_tokens",
    "pid_props_total",
    "pid_props_unique",
    "pid_density_total",
    "pid_density_unique",
    "n_pics_used",
    "graph_transitions_total",
    "graph_edges_total_occ",  # alias
    "graph_RE_total",
    "graph_RE_per100",
    "graph_L1_total",
    "graph_L1_per100",
    "graph_L2_total",
    "graph_L2_per100",
    "graph_edge_entropy",
]

# Variables tracked in Missingness sheet (order matters)
MISSINGNESS_VARS = [
    "participant_id","group","sex","age","education","moca_total",
    "lii_mean","silence_to_audio_ratio","avg_branching_factor","mls_tokens",
    "tree_depth_mean","dependency_length_mean","noun_to_pronoun_ratio","noun_to_verb_ratio",
    "graph_density_noloop","graph_lscc_ratio","graph_self_loops",
    "coh_local_mean","coh_global_mean","coh_composite_mean",
    "pid_repeat_ratio","graph_repeated_edges_per100",
    "graph_tokens","graph_nodes","graph_edges_unique","graph_edges_unique_noloop",
    "graph_lscc_size","graph_lcc_size","graph_lcc_ratio","graph_avg_degree",
    "coh_n_utts","coh_local_sd","coh_global_sd",
    "pid_tokens","pid_props_total","pid_props_unique","pid_density_total","pid_density_unique",
    "n_pics_used",
]


# ----------------- HELPERS -----------------
def holm_adjust(pvals: np.ndarray) -> np.ndarray:
    """Holm step-down adjustment (matches the workbook’s behavior)."""
    pvals = np.asarray(pvals, dtype=float)
    n = len(pvals)
    order = np.argsort(pvals)
    adj = np.empty(n, dtype=float)

    for i, idx in enumerate(order):
        adj[idx] = min(1.0, (n - i) * pvals[idx])

    # enforce monotonicity in sorted order
    adj_sorted = adj[order]
    for i in range(1, n):
        if adj_sorted[i] < adj_sorted[i - 1]:
            adj_sorted[i] = adj_sorted[i - 1]
    adj[order] = adj_sorted
    return adj


def write_df_to_sheet(wb: openpyxl.Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    """Create/replace a sheet and write a dataframe into it."""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append([None if (isinstance(x, float) and np.isnan(x)) else x for x in row])


def compute_missingness(df: pd.DataFrame, variables: list[str]) -> pd.DataFrame:
    n_total = len(df)
    rows = []

    for var in variables:
        row = {"variable": var, "n_total": n_total}
        row["n_missing"] = int(df[var].isna().sum()) if var in df.columns else n_total
        row["pct_missing"] = float(row["n_missing"] / n_total * 100) if n_total else np.nan

        # group Ns
        for g in GROUP_ORDER:
            row[f"{g}_n"] = int((df["group"] == g).sum())

        # group missing
        for g in GROUP_ORDER:
            sub = df[df["group"] == g]
            miss = int(sub[var].isna().sum()) if var in sub.columns else len(sub)
            row[f"{g}_missing"] = miss

        # group pct missing
        for g in GROUP_ORDER:
            sub = df[df["group"] == g]
            n = len(sub)
            miss = int(sub[var].isna().sum()) if var in sub.columns else n
            row[f"{g}_pct_missing"] = float(miss / n * 100) if n else np.nan

        rows.append(row)

    cols = [
        "variable", "n_total", "n_missing", "pct_missing",
        "Control_n", "MCI_n", "Mild_AD_n",
        "Control_missing", "MCI_missing", "Mild_AD_missing",
        "Control_pct_missing", "MCI_pct_missing", "Mild_AD_pct_missing",
    ]
    return pd.DataFrame(rows)[cols]


def desc_demographics(df: pd.DataFrame) -> pd.DataFrame:
    vars_ = ["age", "education", "moca_total"]
    rows = []
    for g in GROUP_ORDER:
        sub = df[df["group"] == g]
        for var in vars_:
            x = sub[var].dropna()
            rows.append({
                "group": g,
                "variable": var,
                "n_nonmissing": int(len(x)),
                "mean": float(x.mean()) if len(x) else np.nan,
                "sd": float(x.std(ddof=1)) if len(x) > 1 else (0.0 if len(x) == 1 else np.nan),
                "median": float(x.median()) if len(x) else np.nan,
                "q1": float(x.quantile(0.25)) if len(x) else np.nan,
                "q3": float(x.quantile(0.75)) if len(x) else np.nan,
                "min": float(x.min()) if len(x) else np.nan,
                "max": float(x.max()) if len(x) else np.nan,
            })
    return pd.DataFrame(rows)


def sex_distribution(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for g in GROUP_ORDER:
        sub = df[df["group"] == g]
        counts = sub["sex"].value_counts(dropna=False)
        for sex, n in counts.items():
            rows.append({"group": g, "sex": sex, "n": int(n)})
    return pd.DataFrame(rows)


def desc_language(df: pd.DataFrame, variables: list[str]) -> pd.DataFrame:
    rows = []
    for g in GROUP_ORDER + [ALL_LABEL]:
        sub = df if g == ALL_LABEL else df[df["group"] == g]
        for var in variables:
            x = sub[var].dropna()
            rows.append({
                "group": g,
                "variable": var,
                "n": int(len(x)),
                "mean": float(x.mean()) if len(x) else np.nan,
                "sd": float(x.std(ddof=1)) if len(x) > 1 else (0.0 if len(x) == 1 else np.nan),
                "min": float(x.min()) if len(x) else np.nan,
                "max": float(x.max()) if len(x) else np.nan,
            })
    return pd.DataFrame(rows)


def ancova_all(df: pd.DataFrame, outcomes: list[str]) -> pd.DataFrame:
    rows = []
    eff_order = [
        "C(group, Treatment(reference='Control'))",
        "C(sex)",
        "age",
        "education",
        "lii_mean",
        "Residual",
    ]

    for outcome in outcomes:
        for model_name in ["ModelA_no_LII", "ModelB_with_LII"]:
            if model_name == "ModelB_with_LII" and outcome == "lii_mean":
                continue

            formula = (
                f"{outcome} ~ C(group, Treatment(reference='Control')) + age + education + C(sex)"
            )
            if model_name == "ModelB_with_LII":
                formula += " + lii_mean"

            model = smf.ols(formula, data=df).fit()
            tab = anova_lm(model, typ=2).reset_index().rename(columns={"index": "effect", "PR(>F)": "p"})
            tab["outcome"] = outcome
            tab["model"] = model_name
            tab["mean_sq"] = tab["sum_sq"] / tab["df"]
            tab["n_used"] = int(model.nobs)

            tab = tab[["outcome","model","effect","df","sum_sq","mean_sq","F","p","n_used"]]
            rows.append(tab)

    anc = pd.concat(rows, ignore_index=True)

    anc["outcome"] = pd.Categorical(anc["outcome"], categories=outcomes, ordered=True)
    anc["model"] = pd.Categorical(anc["model"], categories=["ModelA_no_LII","ModelB_with_LII"], ordered=True)
    anc["effect"] = pd.Categorical(anc["effect"], categories=eff_order, ordered=True)

    anc = anc.sort_values(["outcome","model","effect"]).reset_index(drop=True)
    anc["df"] = anc["df"].astype(int)
    return anc


def posthoc_all(df: pd.DataFrame, outcomes: list[str]) -> pd.DataFrame:
    contrast_order = ["MCI_vs_Control", "Mild_AD_vs_Control", "Mild_AD_vs_MCI"]
    rows = []

    for outcome in outcomes:
        for model_name in ["ModelA_no_LII", "ModelB_with_LII"]:
            if model_name == "ModelB_with_LII" and outcome == "lii_mean":
                continue

            formula = (
                f"{outcome} ~ C(group, Treatment(reference='Control')) + age + education + C(sex)"
            )
            if model_name == "ModelB_with_LII":
                formula += " + lii_mean"

            model = smf.ols(formula, data=df).fit()
            params = model.params.index.tolist()
            df_res = float(model.df_resid)
            tcrit = float(tdist.ppf(0.975, df_res))

            def group_vec(level: str) -> np.ndarray:
                v = np.zeros(len(params), dtype=float)
                if level == "Control":
                    return v
                term = f"C(group, Treatment(reference='Control'))[T.{level}]"
                if term in params:
                    v[params.index(term)] = 1.0
                return v

            contrasts = [("MCI","Control"), ("Mild_AD","Control"), ("Mild_AD","MCI")]
            tmp = []
            for a,b in contrasts:
                L = group_vec(a) - group_vec(b)
                ttest = model.t_test(L)

                est = float(ttest.effect)
                se  = float(ttest.sd)
                tval= float(ttest.tvalue)
                p   = float(ttest.pvalue)
                ci_low  = est - tcrit * se
                ci_high = est + tcrit * se

                tmp.append({
                    "outcome": outcome,
                    "model": model_name,
                    "contrast": f"{a}_vs_{b}",
                    "estimate": est,
                    "se": se,
                    "t": tval,
                    "df2": df_res,
                    "p_raw": p,
                    "ci95_low": ci_low,
                    "ci95_high": ci_high,
                    "n_used": int(model.nobs),
                })

            tmp_df = pd.DataFrame(tmp)
            tmp_df["p_holm"] = holm_adjust(tmp_df["p_raw"].values)
            tmp_df["sig_holm_0_05"] = tmp_df["p_holm"] < 0.05
            rows.append(tmp_df)

    post = pd.concat(rows, ignore_index=True)
    post["outcome"] = pd.Categorical(post["outcome"], categories=outcomes, ordered=True)
    post["model"] = pd.Categorical(post["model"], categories=["ModelA_no_LII","ModelB_with_LII"], ordered=True)
    post["contrast"] = pd.Categorical(post["contrast"], categories=contrast_order, ordered=True)
    post = post.sort_values(["outcome","model","contrast"]).reset_index(drop=True)
    return post


# ----------------- MAIN -----------------
def main():
    # Read unified dataset
    df = pd.read_excel(IN_PATH, sheet_name=SHEET_UNIFIED)

    # Add composite coherence metric
    df["coh_composite_mean"] = df[["coh_local_mean","coh_global_mean"]].mean(axis=1, skipna=True)

    # Recompute sheets
    miss = compute_missingness(df, MISSINGNESS_VARS)
    dem  = desc_demographics(df)
    sex  = sex_distribution(df)
    desc = desc_language(df, OUTCOME_ORDER)
    anc  = ancova_all(df, OUTCOME_ORDER)
    post = posthoc_all(df, OUTCOME_ORDER)

    # Write all sheets
    wb = openpyxl.load_workbook(IN_PATH)

    # Keep MergeDiagnostics as-is if present; otherwise skip/create empty
    if SHEET_DIAG not in wb.sheetnames:
        write_df_to_sheet(wb, SHEET_DIAG, pd.DataFrame({"note": ["(not available in input workbook)"]}))

    write_df_to_sheet(wb, SHEET_UNIFIED, df)
    write_df_to_sheet(wb, SHEET_MISS, miss)
    write_df_to_sheet(wb, SHEET_DEM, dem)
    write_df_to_sheet(wb, SHEET_SEX, sex)
    write_df_to_sheet(wb, SHEET_DESC, desc)
    write_df_to_sheet(wb, SHEET_ANCOVA, anc)
    write_df_to_sheet(wb, SHEET_POSTHOC, post)

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")


if __name__ == "__main__":
    main()
